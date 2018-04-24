import openpyxl
import os
from tkinter import *
from tkinter import ttk

#working directory C:\Users\sarveli\Desktop\VSRP_WorkingSheet Rev 03.xlsx    4 4 53 14692
os.chdir('C:\\Users\\sarveli\\Desktop')
def showBeginsHere():
    #get excel file path
    inputFilePath = inputDataFilePath.get()
    #opening xl workbook
    nwr = openpyxl.load_workbook(inputFilePath)
    #assigning and opening xl worksheet
    sheet1 = nwr.get_sheet_by_name('Final')
    #creating and opening txt file
    text = "sampleText.txt"
    saveFile = open(text, "w")
    #taking video length from text block of GUI
    hh = int(hours.get())
    mm = int(minutes.get())
    ss = int(seconds.get())
    #taking gps cordinate from text block of GUI
    #gpsCounts = int(totalFrame.get())
    #converting time into sec
    hhs = int(hh) * 3600
    mms = int(mm) * 60
    videoLength = hhs + mms + ss
    #divide content of subtitles equally
    frame = videoLength
    #flag denotes iteration of variable videoLength in while loop
    flag = 0.0
    #f1, f2 and f3 are use to hold previous time value
    f1 = 0
    f2 = 0
    f3 = 0
    #ft holds the value of frame count in srt file
    ft = 1
    #i to point current row and columns in xl
    i = 1
    #shows status of program
    status = " Program Started "
    #print(status)
    #writing txt file copying value of xl
    while ( flag < videoLength):
        #print(i)
        #chainage
        x = sheet1.cell(row=i, column=1).value
        #Easting
        y = sheet1.cell(row=i, column=2).value
        #Northing
        z = sheet1.cell(row=i, column=3).value
        #village 1
        c = sheet1.cell(row=i, column=4).value
        #village 2
        d = sheet1.cell(row=i, column=5).value
        #Timing
        e = sheet1.cell(row=i, column=7).value
        #if cell is empty break the loop
        if e == None:
            print("breaking becaue of if")
            status = "breaking becaue of if"
            break
        '''
        if x == None:
            x = " "

        if y == None:
            y = " "

        if z == None:
            z = " "

        if c == None:
            c = " "

        if d == None:
            d = " "
        ''' 
        
       
        
        
        #time division accordinf to flag
        t1 = int(flag/3600)
        b = int(flag%3600)
        t2 = int(b/60)
        t3 = int(b%60)
        #writing in txt file .zfill(2) is to make 1--->01 i.e it adds zero automatically
        saveFile.write(str(ft) + "\n" + str(f1).zfill(2) +":"+ str(f2).zfill(2)+":"+
                       str(f3).zfill(2)+" --> "+str(t1).zfill(2)+":"+str(t2).zfill(2)+":"+str(t3).zfill(2))

        if x == None:
            saveFile.write("\n" )

        if x != None:
            x = round(x)
            saveFile.write("\nCH : "+str(x))
            #print("d=="+str(d))
        
        if y != None and z != None:
            y = round(y , 2)
            z = round(z , 2)
            saveFile.write("; E : " + str(y) + "; N : "+ str(z))
            #print("d=="+str(d))
            

        if c != None or d != None:
            if c != None:
                saveFile.write("\nVillage :  "+str(c))
            if d != None:
                if c == None:
                    saveFile.write("\nVillage :  "+str(d)+"\n")
                else:
                    saveFile.write("; "+str(d)+"\n")

                
            saveFile.write("\n")

        
        saveFile.write("\n")
        #assigning previous values
        f1 = t1
        f2 = t2
        f3 = t3
        #incrementing frame count
        ft = ft +1
        #incrementing time
        flag = flag + 1 
        #pointing towards next cell in xl
        i = i + 1 
    status = " Program End "
    #print(status)
    saveFile.close()

#call TK() function for creating GUI
root = Tk()

#assigning variables of storing user inputs i.e xl path, total no. of frames and duration of video in HH:MM:SS format
hours = StringVar()
minutes = StringVar()
seconds = StringVar()
totalFrame = StringVar()
inputDataFilePath = StringVar()

#giving title and dimension to GUI
root.wm_title("Now_at_Pigeon_Innovative")
root.wm_geometry("500x500")

#creating a frame
frame = Frame(root)
frame.pack(side = TOP)

"""

File Pat                 [________________]
Enter Hours              [________________]
Enter Minutes            [________________]
Enter Seconds            [________________]
Duration of Video (s)    [________________] 
Enter Total nos. of Frame[________________]
Status                   [________________]


                [ OK ]
 
"""
labelDataFilePath = Label(frame, text = "File Path", width =30)
labelDataFilePath.grid(row = 0, column = 0, sticky = E)
entryDataFilePath = Entry(frame, textvariable = inputDataFilePath, width = 50)
entryDataFilePath.grid(row = 0, column = 1, sticky = E)

labelH = Label(frame, text = "Enter Hours", width =30)
labelH.grid(row = 1, column = 0, sticky = E)
entryH = Entry(frame, textvariable = hours, width = 50)
entryH.grid(row = 1, column = 1, sticky = E)

labelM = Label(frame, text = "Enter Minutes", width =30)
labelM.grid(row = 2, column = 0, sticky = E)
entryM = Entry(frame, textvariable = minutes, width = 50)
entryM.grid(row = 2, column = 1, sticky = E)

labelS = Label(frame, text = "Enter Seconds", width =30)
labelS.grid(row = 3, column = 0, sticky = E)
entryS = Entry(frame, textvariable = seconds, width = 50)
entryS.grid(row = 3, column = 1, sticky = E)
"""
labelVidDuration = Label(frame, text = "Duration of Video (s)")
labelVidDuration.grid(row = 4, column = 0, sticky = E)
entryVidDuration = Entry(frame, textvariable = videoLength, width = 50)
entryVidDuration.grid(row = 4, column = 1, sticky = E)

labelTotalFrame = Label(frame, text = "Enter Total nos. of Frame", width =30)
labelTotalFrame.grid(row = 5, column = 0, sticky = E)
entryTotalFrame = Entry(frame, textvariable = totalFrame, width = 50)
entryTotalFrame.grid(row = 5, column = 1, sticky = E)

labelStatus = Label(frame, text = "Status")
labelStatus.grid(row = 6, column = 0, sticky = E)
entryStatus = Entry(frame, textvariable = status, width = 50)
entryStatus.grid(row = 6, column = 1, sticky = E)
"""
okButton = Button(frame, text = "ok", command = showBeginsHere)
okButton.grid(columnspan = 2)

root.mainloop()
